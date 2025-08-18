import streamlit as st
import pandas as pd
import os
import time
import psutil
import gc
import json
import hashlib
from datetime import datetime, timedelta
from youtube_crawler import YouTubeCrawler
from typing import Dict, List, Optional, Any
import logging

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 키워드 분석 함수
def perform_keyword_analysis(texts: List[str]) -> Dict[str, Any]:
    """댓글 텍스트에서 키워드 분석 수행"""
    try:
        import re
        from collections import Counter
        from textblob import TextBlob
        
        # 모든 텍스트를 하나로 합치기
        combined_text = ' '.join(texts)
        
        # 텍스트 전처리
        # 특수문자 제거, 소문자 변환
        cleaned_text = re.sub(r'[^\w\s]', '', combined_text.lower())
        
        # 한국어와 영어 단어 분리
        words = []
        for text in texts:
            # 한국어 단어 추출
            korean_words = re.findall(r'[가-힣]+', text)
            words.extend(korean_words)
            
            # 영어 단어 추출
            english_words = re.findall(r'\b[a-zA-Z]+\b', text.lower())
            words.extend(english_words)
        
        # 불용어 제거
        stop_words = {
            '이', '그', '저', '것', '수', '등', '때', '곳', '말', '일', '년', '월', '일',
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of',
            'with', 'by', 'is', 'are', 'was', 'were', 'be', 'been', 'have', 'has',
            'had', 'do', 'does', 'did', 'will', 'would', 'could', 'should', 'may',
            'might', 'can', 'must', 'shall', 'am', 'i', 'you', 'he', 'she', 'it',
            'we', 'they', 'me', 'him', 'her', 'us', 'them', 'my', 'your', 'his',
            'her', 'its', 'our', 'their', 'mine', 'yours', 'hers', 'ours', 'theirs'
        }
        
        filtered_words = [word for word in words if word not in stop_words and len(word) > 1]
        
        # 단어 빈도 계산
        word_counts = Counter(filtered_words)
        top_keywords = word_counts.most_common(20)
        
        # 키워드 통계
        total_keywords = len(filtered_words)
        unique_keywords = len(word_counts)
        avg_length = sum(len(word) for word in filtered_words) / total_keywords if total_keywords > 0 else 0
        
        # 감정 분석
        sentiment_scores = []
        for text in texts:
            try:
                blob = TextBlob(text)
                sentiment_scores.append(blob.sentiment.polarity)
            except:
                sentiment_scores.append(0)
        
        # 감정 분류
        positive_count = sum(1 for score in sentiment_scores if score > 0.1)
        negative_count = sum(1 for score in sentiment_scores if score < -0.1)
        neutral_count = len(sentiment_scores) - positive_count - negative_count
        
        total_sentiments = len(sentiment_scores)
        sentiment_analysis = {
            'positive': (positive_count / total_sentiments * 100) if total_sentiments > 0 else 0,
            'negative': (negative_count / total_sentiments * 100) if total_sentiments > 0 else 0,
            'neutral': (neutral_count / total_sentiments * 100) if total_sentiments > 0 else 0
        }
        
        return {
            'top_keywords': top_keywords,
            'keyword_stats': {
                'total_keywords': total_keywords,
                'unique_keywords': unique_keywords,
                'avg_length': avg_length
            },
            'sentiment_analysis': sentiment_analysis,
            'processed_texts': len(texts)
        }
        
    except ImportError:
        # TextBlob이 없는 경우 기본 분석
        logger.warning("TextBlob이 설치되지 않아 기본 키워드 분석만 수행합니다.")
        return perform_basic_keyword_analysis(texts)
    except Exception as e:
        logger.error(f"키워드 분석 중 오류: {e}")
        return {
            'top_keywords': [],
            'keyword_stats': {'total_keywords': 0, 'unique_keywords': 0, 'avg_length': 0},
            'sentiment_analysis': {'positive': 0, 'negative': 0, 'neutral': 0},
            'processed_texts': 0
        }

def perform_basic_keyword_analysis(texts: List[str]) -> Dict[str, Any]:
    """기본 키워드 분석 (TextBlob 없이)"""
    try:
        import re
        from collections import Counter
        
        # 모든 텍스트를 하나로 합치기
        combined_text = ' '.join(texts)
        
        # 텍스트 전처리
        cleaned_text = re.sub(r'[^\w\s]', '', combined_text.lower())
        
        # 한국어와 영어 단어 분리
        words = []
        for text in texts:
            # 한국어 단어 추출
            korean_words = re.findall(r'[가-힣]+', text)
            words.extend(korean_words)
            
            # 영어 단어 추출
            english_words = re.findall(r'\b[a-zA-Z]+\b', text.lower())
            words.extend(english_words)
        
        # 불용어 제거
        stop_words = {
            '이', '그', '저', '것', '수', '등', '때', '곳', '말', '일', '년', '월', '일',
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of'
        }
        
        filtered_words = [word for word in words if word not in stop_words and len(word) > 1]
        
        # 단어 빈도 계산
        word_counts = Counter(filtered_words)
        top_keywords = word_counts.most_common(20)
        
        # 키워드 통계
        total_keywords = len(filtered_words)
        unique_keywords = len(word_counts)
        avg_length = sum(len(word) for word in filtered_words) / total_keywords if total_keywords > 0 else 0
        
        return {
            'top_keywords': top_keywords,
            'keyword_stats': {
                'total_keywords': total_keywords,
                'unique_keywords': unique_keywords,
                'avg_length': avg_length
            },
            'sentiment_analysis': {'positive': 0, 'negative': 0, 'neutral': 0},
            'processed_texts': len(texts)
        }
        
    except Exception as e:
        logger.error(f"기본 키워드 분석 중 오류: {e}")
        return {
            'top_keywords': [],
            'keyword_stats': {'total_keywords': 0, 'unique_keywords': 0, 'avg_length': 0},
            'sentiment_analysis': {'positive': 0, 'negative': 0, 'neutral': 0},
            'processed_texts': 0
        }

# plotly 대신 streamlit의 기본 차트 기능 사용
PLOTLY_AVAILABLE = False

# 성능 모니터링 클래스
class PerformanceMonitor:
    def __init__(self):
        self.start_time = None
        self.memory_usage = []
        self.cpu_usage = []
        self.operation_times = {}
    
    def start_monitoring(self):
        self.start_time = time.time()
        self.memory_usage = []
        self.cpu_usage = []
        self.operation_times = {}
    
    def record_operation(self, operation_name: str, duration: float):
        if operation_name not in self.operation_times:
            self.operation_times[operation_name] = []
        self.operation_times[operation_name].append(duration)
    
    def get_memory_usage(self):
        process = psutil.Process()
        return process.memory_info().rss / 1024 / 1024  # MB
    
    def get_cpu_usage(self):
        return psutil.cpu_percent()
    
    def get_performance_summary(self):
        if not self.start_time:
            return {}
        
        total_time = time.time() - self.start_time
        avg_memory = sum(self.memory_usage) / len(self.memory_usage) if self.memory_usage else 0
        avg_cpu = sum(self.cpu_usage) / len(self.cpu_usage) if self.cpu_usage else 0
        
        return {
            'total_time': total_time,
            'avg_memory_mb': avg_memory,
            'avg_cpu_percent': avg_cpu,
            'operations': self.operation_times
        }

# 히스토리 관리 클래스
class HistoryManager:
    def __init__(self):
        self.history_file = "download_history.json"
        self.max_history = 50
    
    def add_download_record(self, filename: str, data_type: str, record_count: int, 
                           file_size: int, download_time: str):
        history = self.load_history()
        
        record = {
            'id': hashlib.md5(f"{filename}{download_time}".encode()).hexdigest()[:8],
            'filename': filename,
            'data_type': data_type,
            'record_count': record_count,
            'file_size_mb': round(file_size / 1024 / 1024, 2),
            'download_time': download_time,
            'timestamp': datetime.now().isoformat()
        }
        
        history.insert(0, record)
        
        # 최대 기록 수 제한
        if len(history) > self.max_history:
            history = history[:self.max_history]
        
        self.save_history(history)
        return record
    
    def load_history(self) -> List[Dict]:
        try:
            if os.path.exists(self.history_file):
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"히스토리 로드 오류: {e}")
        return []
    
    def save_history(self, history: List[Dict]):
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"히스토리 저장 오류: {e}")
    
    def get_recent_history(self, limit: int = 10) -> List[Dict]:
        history = self.load_history()
        return history[:limit]
    
    def clear_history(self):
        try:
            if os.path.exists(self.history_file):
                os.remove(self.history_file)
        except Exception as e:
            logger.error(f"히스토리 삭제 오류: {e}")

# 캐시 관리 클래스
class CacheManager:
    def __init__(self, cache_dir="cache"):
        self.cache_dir = cache_dir
        self.max_cache_size = 200 * 1024 * 1024  # 200MB로 증가
        self.cache_ttl = 24 * 60 * 60  # 24시간 캐시 유효기간
        os.makedirs(cache_dir, exist_ok=True)
        self._cache_stats = {'hits': 0, 'misses': 0, 'saves': 0}
    
    def get_cache_key(self, data: str) -> str:
        """더 정교한 캐시 키 생성"""
        # 데이터와 타임스탬프를 조합하여 키 생성
        timestamp = str(int(time.time() // (60 * 60)))  # 1시간 단위
        return hashlib.md5(f"{data}_{timestamp}".encode()).hexdigest()
    
    def get_cache_path(self, key: str) -> str:
        return os.path.join(self.cache_dir, f"{key}.pkl")
    
    def is_cached(self, key: str) -> bool:
        """캐시 존재 여부 및 유효성 확인"""
        cache_path = self.get_cache_path(key)
        if not os.path.exists(cache_path):
            self._cache_stats['misses'] += 1
            return False
        
        # TTL 확인
        file_age = time.time() - os.path.getmtime(cache_path)
        if file_age > self.cache_ttl:
            try:
                os.remove(cache_path)
                logger.info(f"만료된 캐시 파일 삭제: {key}")
            except:
                pass
            self._cache_stats['misses'] += 1
            return False
        
        self._cache_stats['hits'] += 1
        return True
    
    def save_to_cache(self, key: str, data: Any):
        """캐시 저장 및 압축"""
        try:
            import pickle
            import gzip
            cache_path = self.get_cache_path(key)
            
            # 데이터 압축하여 저장
            with gzip.open(cache_path, 'wb') as f:
                pickle.dump(data, f)
            
            self._cache_stats['saves'] += 1
            self._cleanup_cache()
            logger.info(f"캐시 저장 완료: {key} ({len(str(data))} bytes)")
        except Exception as e:
            logger.error(f"캐시 저장 오류: {e}")
    
    def load_from_cache(self, key: str) -> Optional[Any]:
        """캐시 로드 및 압축 해제"""
        try:
            import pickle
            import gzip
            cache_path = self.get_cache_path(key)
            
            if os.path.exists(cache_path):
                with gzip.open(cache_path, 'rb') as f:
                    data = pickle.load(f)
                logger.info(f"캐시 로드 완료: {key}")
                return data
        except Exception as e:
            logger.error(f"캐시 로드 오류: {e}")
        return None
    
    def get_cache_stats(self) -> Dict[str, int]:
        """캐시 통계 반환"""
        total_requests = self._cache_stats['hits'] + self._cache_stats['misses']
        hit_rate = (self._cache_stats['hits'] / total_requests * 100) if total_requests > 0 else 0
        
        return {
            **self._cache_stats,
            'hit_rate': round(hit_rate, 2),
            'total_requests': total_requests
        }
    
    def _cleanup_cache(self):
        """캐시 크기 제한 및 오래된 파일 정리"""
        try:
            cache_files = []
            for filename in os.listdir(self.cache_dir):
                if filename.endswith('.pkl'):
                    filepath = os.path.join(self.cache_dir, filename)
                    cache_files.append((filepath, os.path.getmtime(filepath)))
            
            # 파일 크기 계산
            total_size = sum(os.path.getsize(f[0]) for f in cache_files)
            
            if total_size > self.max_cache_size:
                # 오래된 파일부터 삭제
                cache_files.sort(key=lambda x: x[1])
                for filepath, _ in cache_files:
                    os.remove(filepath)
                    total_size -= os.path.getsize(filepath)
                    if total_size <= self.max_cache_size * 0.8:  # 80%까지 줄임
                        break
        except Exception as e:
            logger.error(f"캐시 정리 오류: {e}")

# 전역 성능 모니터, 히스토리 매니저, 캐시 매니저 초기화
if 'performance_monitor' not in st.session_state:
    st.session_state.performance_monitor = PerformanceMonitor()

if 'history_manager' not in st.session_state:
    st.session_state.history_manager = HistoryManager()

if 'cache_manager' not in st.session_state:
    st.session_state.cache_manager = CacheManager()

# 페이지 설정
st.set_page_config(
    page_title="유튜브 크롤러",
    page_icon="🎥",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={
        'Get Help': 'https://github.com/your-repo/youtube-crawler',
        'Report a bug': "https://github.com/your-repo/youtube-crawler/issues",
        'About': "# 유튜브 크롤러\n\n현대적인 유튜브 데이터 수집 서비스입니다."
    }
)

# 2025년 디자인 트렌드 CSS 스타일
st.markdown("""
<style>
    /* 전체 페이지 스타일 - 2025년 트렌드 */
    .main {
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
        padding: 0;
        min-height: 100vh;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        color: #1a202c;
    }
    
    /* 스크롤바 스타일 - 미니멀 */
    ::-webkit-scrollbar {
        width: 4px;
    }
    
    ::-webkit-scrollbar-track {
        background: transparent;
    }
    
    ::-webkit-scrollbar-thumb {
        background: rgba(0, 0, 0, 0.2);
        border-radius: 2px;
        backdrop-filter: blur(10px);
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: rgba(0, 0, 0, 0.3);
    }
    
    /* 헤더 스타일 - 2025년 트렌드 */
    .main-header {
        font-size: 3rem;
        font-weight: 700;
        background: linear-gradient(135deg, #1a202c 0%, #4a5568 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        text-align: center;
        margin: 3rem 0 2rem 0;
        letter-spacing: -0.02em;
        line-height: 1.2;
    }
    
    /* 카드 스타일 - 글래스모피즘 */
    .metric-card {
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(20px);
        border-radius: 16px;
        padding: 2rem;
        border: 1px solid rgba(255, 255, 255, 0.2);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        text-align: center;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
    }
    
    .metric-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 20px 40px rgba(0, 0, 0, 0.15);
        background: rgba(255, 255, 255, 0.9);
    }
    
    /* 버튼 스타일 - 네오모피즘 */
    .stButton > button {
        background: linear-gradient(145deg, #ffffff, #e6e6e6);
        border: none;
        border-radius: 12px;
        padding: 1rem 2rem;
        font-weight: 600;
        color: #1a202c;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        box-shadow: 8px 8px 16px rgba(0, 0, 0, 0.1), -8px -8px 16px rgba(255, 255, 255, 0.8);
        width: 100%;
        font-size: 1rem;
        letter-spacing: 0.01em;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 12px 12px 24px rgba(0, 0, 0, 0.15), -12px -12px 24px rgba(255, 255, 255, 0.9);
        background: linear-gradient(145deg, #f8f9fa, #e9ecef);
    }
    
    .stButton > button:active {
        transform: translateY(0);
        box-shadow: inset 4px 4px 8px rgba(0, 0, 0, 0.1), inset -4px -4px 8px rgba(255, 255, 255, 0.8);
    }
    
    /* 사이드바 스타일 - 글래스모피즘 */
    .css-1d391kg {
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(20px);
        border-right: 1px solid rgba(255, 255, 255, 0.2);
        text-align: center;
    }
    
    /* 입력 필드 스타일 - 2025년 트렌드 */
    .stTextInput > div > div > input {
        border-radius: 12px;
        border: 2px solid rgba(0, 0, 0, 0.1);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(10px);
        padding: 1rem 1.25rem;
        font-size: 1rem;
        font-weight: 500;
    }
    
    .stTextInput > div > div > input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1);
        background: rgba(255, 255, 255, 0.95);
        transform: translateY(-1px);
    }
    
    /* 크롤링 진행 과정 스타일 - 인터랙티브 */
    .crawling-progress {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 12px;
        padding: 2rem;
        margin: 1rem 0;
        box-shadow: 0 8px 32px rgba(102, 126, 234, 0.2);
        border: 1px solid rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
    }
    
    .progress-step {
        display: flex;
        align-items: center;
        margin: 1rem 0;
        padding: 1rem;
        background: rgba(255, 255, 255, 0.1);
        border-radius: 8px;
        transition: all 0.3s ease;
        border-left: 4px solid transparent;
    }
    
    .progress-step.active {
        background: rgba(255, 255, 255, 0.2);
        border-left-color: #4ade80;
        transform: translateX(5px);
        box-shadow: 0 4px 12px rgba(74, 222, 128, 0.3);
    }
    
    .progress-step.completed {
        background: rgba(255, 255, 255, 0.15);
        border-left-color: #10b981;
    }
    
    .progress-step.error {
        background: rgba(239, 68, 68, 0.1);
        border-left-color: #ef4444;
    }
    
    .step-icon {
        width: 40px;
        height: 40px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        margin-right: 1rem;
        font-size: 1.2rem;
        font-weight: bold;
        transition: all 0.3s ease;
    }
    
    .step-icon.pending {
        background: rgba(255, 255, 255, 0.2);
        color: rgba(255, 255, 255, 0.7);
    }
    
    .step-icon.active {
        background: #4ade80;
        color: white;
        animation: pulse 2s infinite;
    }
    
    .step-icon.completed {
        background: #10b981;
        color: white;
    }
    
    .step-icon.error {
        background: #ef4444;
        color: white;
    }
    
    .step-content {
        flex: 1;
    }
    
    .step-title {
        font-weight: 600;
        color: white;
        margin-bottom: 0.25rem;
        font-size: 1rem;
    }
    
    .step-description {
        color: rgba(255, 255, 255, 0.8);
        font-size: 0.875rem;
        margin: 0;
    }
    
    .step-details {
        color: rgba(255, 255, 255, 0.6);
        font-size: 0.75rem;
        margin-top: 0.25rem;
    }
    
    /* 애니메이션 */
    @keyframes pulse {
        0% {
            box-shadow: 0 0 0 0 rgba(74, 222, 128, 0.7);
        }
        70% {
            box-shadow: 0 0 0 10px rgba(74, 222, 128, 0);
        }
        100% {
            box-shadow: 0 0 0 0 rgba(74, 222, 128, 0);
        }
    }
    
    @keyframes slideIn {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .progress-step {
        animation: slideIn 0.5s ease-out;
    }
    
    /* 진행률 바 스타일 */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #4ade80, #10b981, #059669);
        border-radius: 10px;
        height: 8px;
    }
    
    .stProgress > div > div > div {
        background: rgba(255, 255, 255, 0.2);
        border-radius: 10px;
        height: 8px;
    }
    
    /* 상태 메시지 스타일 */
    .status-message {
        background: rgba(255, 255, 255, 0.1);
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
        border-left: 4px solid #3182ce;
        animation: slideIn 0.3s ease-out;
    }
    
    .status-message.success {
        border-left-color: #10b981;
        background: rgba(16, 185, 129, 0.1);
    }
    
    .status-message.error {
        border-left-color: #ef4444;
        background: rgba(239, 68, 68, 0.1);
    }
    
    .status-message.warning {
        border-left-color: #f59e0b;
        background: rgba(245, 158, 11, 0.1);
    }
    
    .status-message.info {
        border-left-color: #3182ce;
        background: rgba(49, 130, 206, 0.1);
    }
    
    /* 로딩 스피너 스타일 */
    .loading-spinner {
        display: inline-block;
        width: 20px;
        height: 20px;
        border: 3px solid rgba(255, 255, 255, 0.3);
        border-radius: 50%;
        border-top-color: #fff;
        animation: spin 1s ease-in-out infinite;
        margin-right: 0.5rem;
    }
    
    @keyframes spin {
        to { transform: rotate(360deg); }
    }
    
    /* 실시간 업데이트 카운터 */
    .live-counter {
        background: rgba(255, 255, 255, 0.1);
        border-radius: 6px;
        padding: 0.5rem 1rem;
        display: inline-block;
        margin: 0.25rem;
        font-weight: 600;
        color: white;
        animation: slideIn 0.3s ease-out;
    }
    
    .live-counter.primary {
        background: rgba(49, 130, 206, 0.2);
        border: 1px solid rgba(49, 130, 206, 0.3);
    }
    
    .live-counter.success {
        background: rgba(16, 185, 129, 0.2);
        border: 1px solid rgba(16, 185, 129, 0.3);
    }
    
    .live-counter.warning {
        background: rgba(245, 158, 11, 0.2);
        border: 1px solid rgba(245, 158, 11, 0.3);
    }
    }
    
    /* 텍스트 영역 스타일 - 2025년 트렌드 */
    .stTextArea > div > div > textarea {
        border-radius: 12px;
        border: 2px solid rgba(0, 0, 0, 0.1);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(10px);
        padding: 1rem 1.25rem;
        font-size: 1rem;
        font-weight: 500;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    }
    
    .stTextArea > div > div > textarea:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1);
        background: rgba(255, 255, 255, 0.95);
        transform: translateY(-1px);
    }
    
    /* 숫자 입력 필드 스타일 - 2025년 트렌드 */
    .stNumberInput > div > div > input {
        border-radius: 12px;
        border: 2px solid rgba(0, 0, 0, 0.1);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(10px);
        padding: 1rem 1.25rem;
        font-size: 1rem;
        font-weight: 500;
    }
    
    .stNumberInput > div > div > input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1);
        background: rgba(255, 255, 255, 0.95);
        transform: translateY(-1px);
    }
    
    /* 날짜 입력 필드 스타일 - 2025년 트렌드 */
    .stDateInput > div > div > input {
        border-radius: 12px;
        border: 2px solid rgba(0, 0, 0, 0.1);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(10px);
        padding: 1rem 1.25rem;
        font-size: 1rem;
        font-weight: 500;
    }
    
    .stDateInput > div > div > input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1);
        background: rgba(255, 255, 255, 0.95);
        transform: translateY(-1px);
    }
    
    /* 체크박스 스타일 - 2025년 트렌드 */
    .stCheckbox > div > div > div {
        border-radius: 8px;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    /* 성공 메시지 - 2025년 트렌드 */
    .success-message {
        background: rgba(56, 161, 105, 0.1);
        color: #1a202c;
        padding: 1.25rem;
        border-radius: 12px;
        border: 1px solid rgba(56, 161, 105, 0.2);
        font-weight: 600;
        backdrop-filter: blur(10px);
    }
    
    /* 정보 메시지 - 2025년 트렌드 */
    .info-message {
        background: rgba(102, 126, 234, 0.1);
        color: #1a202c;
        padding: 1.25rem;
        border-radius: 12px;
        border: 1px solid rgba(102, 126, 234, 0.2);
        font-weight: 600;
        backdrop-filter: blur(10px);
    }
    
    /* 경고 메시지 - 2025년 트렌드 */
    .warning-message {
        background: rgba(229, 62, 62, 0.1);
        color: #1a202c;
        padding: 1.25rem;
        border-radius: 12px;
        border: 1px solid rgba(229, 62, 62, 0.2);
        font-weight: 600;
        backdrop-filter: blur(10px);
    }
    
    /* 진행률 바 스타일 - 2025년 트렌드 */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        border-radius: 8px;
        box-shadow: 0 2px 8px rgba(102, 126, 234, 0.3);
    }
    
    /* 탭 스타일 - 2025년 트렌드 */
    .stTabs > div > div > div > div {
        border-radius: 12px 12px 0 0;
        border: 1px solid rgba(255, 255, 255, 0.2);
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(10px);
    }
    
    /* 데이터프레임 스타일 - 2025년 트렌드 */
    .dataframe {
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
        background: rgba(255, 255, 255, 0.9);
        backdrop-filter: blur(20px);
    }
    
    /* 애니메이션 - 2025년 트렌드 */
    @keyframes fadeIn {
        from { 
            opacity: 0; 
            transform: translateY(30px) scale(0.95); 
        }
        to { 
            opacity: 1; 
            transform: translateY(0) scale(1); 
        }
    }
    
    @keyframes slideIn {
        from { 
            opacity: 0; 
            transform: translateX(-40px) scale(0.95); 
        }
        to { 
            opacity: 1; 
            transform: translateX(0) scale(1); 
        }
    }
    
    @keyframes pulse {
        0% { transform: scale(1); }
        50% { transform: scale(1.02); }
        100% { transform: scale(1); }
    }
    
    .fade-in {
        animation: fadeIn 0.8s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    .slide-in {
        animation: slideIn 0.6s cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    .pulse {
        animation: pulse 3s infinite cubic-bezier(0.4, 0, 0.2, 1);
    }
    
    /* 로딩 스피너 - 2025년 트렌드 */
    .loading-spinner {
        display: inline-block;
        width: 24px;
        height: 24px;
        border: 3px solid rgba(102, 126, 234, 0.2);
        border-radius: 50%;
        border-top-color: #667eea;
        animation: spin 1.2s cubic-bezier(0.4, 0, 0.2, 1) infinite;
    }
    
    @keyframes spin {
        to { transform: rotate(360deg); }
    }
    

    
    /* 반응형 디자인 - 2025년 트렌드 */
    @media (max-width: 768px) {
        .main-header {
            font-size: 2.5rem;
            margin: 2rem 0 1.5rem 0;
        }
        
        .metric-card {
            padding: 1.5rem;
            border-radius: 12px;
        }
        
        .stButton > button {
            padding: 0.875rem 1.5rem;
            font-size: 0.95rem;
        }
        
        .stTextInput > div > div > input,
        .stNumberInput > div > div > input,
        .stDateInput > div > div > input {
            padding: 0.875rem 1rem;
            font-size: 0.95rem;
        }
    }
    
    @media (max-width: 480px) {
        .main-header {
            font-size: 2rem;
            margin: 1.5rem 0 1rem 0;
        }
        
        .metric-card {
            padding: 1rem;
            border-radius: 10px;
        }
        
        .stButton > button {
            padding: 0.75rem 1.25rem;
            font-size: 0.9rem;
        }
    }
</style>
""", unsafe_allow_html=True)

def main():
    # 시스템 점검 및 성능 모니터링 시작
    performance_monitor = st.session_state.performance_monitor
    history_manager = st.session_state.history_manager
    cache_manager = st.session_state.cache_manager
    
    # 성능 모니터링 시작
    performance_monitor.start_monitoring()
    
    # 2025년 트렌드 헤더
    st.markdown('<h1 class="main-header fade-in">유튜브 크롤러</h1>', unsafe_allow_html=True)
    
    # 서브타이틀
    st.markdown('<p style="text-align: center; color: #4a5568; font-size: 1.1rem; margin-bottom: 2rem; font-weight: 400;">유튜브 데이터 수집 및 분석 서비스(since 2025)</p>', unsafe_allow_html=True)
    
    # 서비스 탭 선택
    tab1, tab2 = st.tabs(["🎯 키워드 기반 크롤링", "💬 영상 ID 댓글 추출"])
    
    # 시스템 상태 표시
    with st.sidebar:
        st.markdown("### 🔧 시스템 상태")
        
        # 메모리 사용량
        memory_usage = performance_monitor.get_memory_usage()
        st.metric("메모리 사용량", f"{memory_usage:.1f} MB")
        
        # CPU 사용량
        cpu_usage = performance_monitor.get_cpu_usage()
        st.metric("CPU 사용량", f"{cpu_usage:.1f}%")
        
        # 캐시 상태
        cache_files = len([f for f in os.listdir(cache_manager.cache_dir) if f.endswith('.pkl')])
        st.metric("캐시 파일 수", cache_files)
        
        # 성능 최적화 버튼
        if st.button("🧹 메모리 정리"):
            # 메모리 정리 전 상태
            before_memory = performance_monitor.get_memory_usage()
            
            # 가비지 컬렉션 실행
            collected = gc.collect()
            
            # 캐시 정리
            cache_manager._cleanup_cache()
            
            # 메모리 정리 후 상태
            after_memory = performance_monitor.get_memory_usage()
            memory_freed = before_memory - after_memory
            
            st.success(f"메모리 정리 완료! {memory_freed:.1f}MB 해제됨 (수집된 객체: {collected}개)")
            st.rerun()
        
        # 캐시 통계 표시
        cache_stats = cache_manager.get_cache_stats()
        st.markdown("### 📊 캐시 통계")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("캐시 히트율", f"{cache_stats.get('hit_rate', 0)}%")
        with col2:
            st.metric("캐시 저장", cache_stats.get('saves', 0))
        with col3:
            st.metric("캐시 요청", cache_stats.get('total_requests', 0))
    
    # 실시간 알림 표시 (현재 비활성화)
    
    # 크롤링 진행 중 표시 (세션 상태 확인)
    if hasattr(st.session_state, 'crawling_completed') and not st.session_state.crawling_completed:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 1.5rem; border-radius: 12px; margin: 1rem 0; text-align: center;">
            <h3 style="margin: 0 0 0.5rem 0;">🔄 크롤링이 진행 중입니다</h3>
            <p style="margin: 0; opacity: 0.9;">실제 크롤링 및 수집은 계속 진행중입니다. 페이지를 닫아도 백그라운드에서 작업이 계속됩니다.</p>
        </div>
        """, unsafe_allow_html=True)
    
    # 첫 번째 탭: 키워드 기반 크롤링
    with tab1:
        # 통합 레이아웃 - 상단에 설정, 하단에 크롤링과 분석을 나란히 배치
        with st.container():
            # 상단 설정 영역
            st.markdown('<h2 style="color: #1a202c; font-size: 1.5rem; font-weight: 600; margin-bottom: 1rem;">키워드 기반 크롤링 설정</h2>', unsafe_allow_html=True)
        
        # 설정을 3개 컬럼으로 배치
        col1, col2, col3 = st.columns([1, 1, 1])
        
        with col1:
            st.markdown('<h3 style="color: #4a5568; font-size: 1.1rem; font-weight: 500;">검색 설정</h3>', unsafe_allow_html=True)
            
            # 키워드 개수 선택
            keyword_count = st.selectbox(
                "키워드 개수",
                options=[1, 2, 3, 4, 5],
                index=2,
                help="수집할 키워드의 개수를 선택하세요"
            )
            
            # 선택된 개수만큼 키워드 입력 필드 생성
            keywords = []
            for i in range(keyword_count):
                keyword = st.text_input(
                    f"키워드 {i+1}",
                    placeholder=f"키워드 {i+1}을 입력하세요",
                    help=f"검색할 키워드 {i+1}을 입력하세요"
                )
                if keyword.strip():
                    keywords.append(keyword.strip())
            
            # 키워드 검증
            if not keywords:
                st.warning("⚠️ 최소 1개의 키워드를 입력해주세요.")
                st.stop()
        
        with col2:
            st.markdown('<h3 style="color: #4a5568; font-size: 1.1rem; font-weight: 500;">수집 설정</h3>', unsafe_allow_html=True)
            
            videos_per_keyword = st.number_input(
                "키워드당 영상 수",
                min_value=1, max_value=50, value=10,
                step=1,
                help="각 키워드당 수집할 영상의 수"
            )
            
            collect_comments = st.checkbox(
                "💬 댓글 수집",
                value=True,  # 기본값을 True로 변경
                help="영상 댓글 수집 활성화"
            )
            
            if collect_comments:
                # 댓글 수집 최적화 설정
                col1, col2 = st.columns(2)
                with col1:
                    comments_per_video = st.number_input(
                        "영상당 댓글 수",
                        min_value=1, max_value=50, value=5,  # 기본값을 5로 줄임
                        step=1,
                        help="영상당 수집할 댓글의 수 (성능 최적화를 위해 5-10개 권장)"
                    )
                
                with col2:
                    # 댓글 수집 배치 크기 설정
                    comment_batch_size = st.number_input(
                        "댓글 배치 크기",
                        min_value=5, max_value=20, value=10,
                        step=5,
                        help="한 번에 처리할 댓글의 수 (메모리 최적화)"
                    )
                
                # 댓글 수집 최적화 팁
                st.info("💡 **댓글 수집 최적화**:\n"
                       "- 영상당 댓글 수를 5-10개로 제한하면 속도가 빨라집니다\n"
                       "- 배치 크기를 조절하여 메모리 사용량을 제어할 수 있습니다\n"
                       "- 많은 영상을 처리할 때는 댓글 수집을 비활성화하는 것을 고려하세요")
            else:
                comments_per_video = 0
                comment_batch_size = 10
        
        with col3:
            st.markdown('<h3 style="color: #4a5568; font-size: 1.1rem; font-weight: 500;">날짜 & 파일</h3>', unsafe_allow_html=True)
            
            # 날짜 필터링 설정
            use_date_filter = st.checkbox(
                "날짜 범위 설정",
                value=False,
                help="업로드 날짜로 영상 필터링"
            )
            
            start_date = None
            end_date = None
            
            if use_date_filter:
                start_date = st.date_input(
                    "시작 날짜",
                    value=datetime.now() - timedelta(days=30),
                    help="검색 시작 날짜"
                )
                end_date = st.date_input(
                    "종료 날짜",
                    value=datetime.now(),
                    help="검색 종료 날짜"
                )
            
            # 파일명 설정
            filename = st.text_input(
                "출력 파일명",
                value="youtube_data.xlsx",
                help="저장할 엑셀 파일명 (확장자 포함)"
            )
            
            # 고급 설정
            with st.expander("고급 설정"):
                enable_keyword_analysis = st.checkbox(
                    "영상 키워드 분석",
                    value=True,
                    help="댓글에서 키워드 및 감정 분석 수행"
                )
                
                excel_encoding = st.selectbox(
                    "엑셀 인코딩",
                    options=['utf-8-sig', 'utf-8', 'cp949'],
                    index=0,
                    help="엑셀 파일 저장 시 사용할 인코딩"
                )
                
                # 동시 처리 수 최적화
                col1, col2 = st.columns(2)
                with col1:
                    max_workers = st.slider(
                        "동시 처리 수",
                        min_value=1, max_value=8, value=4,
                        help="동시에 처리할 작업의 수"
                    )
                
                with col2:
                    # 시스템 성능에 따른 권장값 표시
                    cpu_count = psutil.cpu_count()
                    recommended_workers = min(cpu_count, 6)  # 최대 6개로 제한
                    
                    if max_workers > recommended_workers:
                        st.warning(f"⚠️ 권장 동시 처리 수: {recommended_workers}개")
                    else:
                        st.success(f"✅ 권장 동시 처리 수: {recommended_workers}개")
                
                # 성능 최적화 팁
                st.info("💡 **성능 최적화 팁**:\n"
                       "- 동시 처리 수는 CPU 코어 수의 50-75%가 적절합니다\n"
                       "- 댓글 수집 시에는 동시 처리 수를 줄이는 것이 안정적입니다\n"
                       "- 메모리 부족 시 동시 처리 수를 줄여보세요")
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
    
    # 구분선
    st.markdown("---")
    
    # 크롤링 실행 버튼 (중앙 배치)
    st.markdown('<div style="text-align: center; margin: 2rem 0;">', unsafe_allow_html=True)
    if st.button("🎯 크롤링 시작", type="primary", use_container_width=False, help="설정된 조건으로 크롤링을 시작합니다"):
        # 크롤링 시작 시 세션 상태 초기화
        st.session_state.crawling_completed = False
        st.session_state.crawling_logs = []
        
        # 성능 모니터링 시작
        performance_monitor.start_monitoring()
        
        # 캐시 키 생성
        cache_key_data = f"{','.join(keywords)}_{videos_per_keyword}_{collect_comments}_{comments_per_video}"
        cache_key = cache_manager.get_cache_key(cache_key_data)
        
        # 캐시된 결과 확인
        if cache_manager.is_cached(cache_key):
            cached_result = cache_manager.load_from_cache(cache_key)
            if cached_result:
                st.success("🚀 캐시된 결과를 불러왔습니다!")
                st.session_state.videos = cached_result.get('videos', [])
                st.session_state.comments = cached_result.get('comments', [])
                st.session_state.crawling_completed = True
                st.rerun()
                return
        
        if not keywords:
            st.error("❌ 키워드를 입력해주세요.")
            st.stop()
        
        # 콤팩트한 진행 상황 표시
        progress_container = st.container()
        
        with progress_container:
            # 진행률 바와 상태 메시지를 한 줄에 표시
            col_progress, col_status = st.columns([3, 1])
            
            with col_progress:
                progress_bar = st.progress(0)
                status_text = st.empty()
            
            with col_status:
                st.markdown("**진행률**")
                progress_text = st.empty()
        
        # 실시간 로그 메시지 저장
        if 'crawling_logs' not in st.session_state:
            st.session_state.crawling_logs = []
        
        def add_log(message, log_type="info"):
            """로그 메시지 추가"""
            timestamp = datetime.now().strftime("%H:%M:%S")
            log_entry = {
                'timestamp': timestamp,
                'message': message,
                'type': log_type
            }
            st.session_state.crawling_logs.append(log_entry)
            
            # 최근 20개 로그만 유지
            if len(st.session_state.crawling_logs) > 20:
                st.session_state.crawling_logs = st.session_state.crawling_logs[-20:]
        
        # 크롤러 실행
        crawler = None
        try:
            # 단계 1: 크롤러 초기화 (콤팩트)
            status_text.text("🔬 크롤러 초기화 중...")
            progress_text.text("10%")
            
            add_log("🔬 크롤러 초기화 시작", "info")
            
            # 설정 적용 (최적화된 버전)
            config = {
                'max_workers': max_workers,
                'enable_keyword_analysis': enable_keyword_analysis,
                'excel_encoding': excel_encoding,
                'max_comments_per_video': comments_per_video if collect_comments else 0,
                'comment_batch_size': comment_batch_size if collect_comments else 20,
                # 네트워크 최적화 설정
                'page_load_timeout': 20,
                'implicit_wait': 5,
                'network_idle_timeout': 3,
                'connection_timeout': 10,
                'request_timeout': 15
            }
            
            crawler = YouTubeCrawler()
            crawler.update_config(config)
            
            add_log("✅ 크롤러 초기화 완료", "success")
            
            progress_bar.progress(0.1)
            
            # 단계 2: 영상 검색 (콤팩트)
            status_text.text("🔍 영상 검색 중...")
            add_log("🔍 영상 검색 시작", "info")
            
            videos = []
            
            for i, keyword in enumerate(keywords):
                progress = 0.1 + (i / len(keywords)) * 0.4  # 10%~50%
                progress_bar.progress(progress)
                
                # 콤팩트한 진행 상태 업데이트
                status_text.text(f"🔍 '{keyword}' 검색 중... ({i+1}/{len(keywords)})")
                progress_text.text(f"{int(progress * 100)}%")
                
                # 날짜 필터링 적용
                start_dt = datetime.combine(start_date, datetime.min.time()) if start_date else None
                end_dt = datetime.combine(end_date, datetime.max.time()) if end_date else None
                
                keyword_videos = crawler.search_videos([keyword], videos_per_keyword, start_dt, end_dt)
                videos.extend(keyword_videos)
            
            if not videos:
                add_log("❌ 검색된 영상이 없습니다.", "error")
                st.error("❌ 검색된 영상이 없습니다.")
                return
            
            # 단계 3: 댓글 수집 (선택사항) - 콤팩트
            all_comments = []
            if collect_comments and videos:
                status_text.text("💬 댓글 수집 중...")
                add_log("💬 댓글 수집 시작", "info")
                
                for i, video in enumerate(videos):
                    progress = 0.5 + (i / len(videos)) * 0.4  # 50%~90%
                    progress_bar.progress(progress)
                    
                    # 콤팩트한 댓글 수집 진행 상태
                    status_text.text(f"💬 댓글 수집 중... ({i+1}/{len(videos)})")
                    progress_text.text(f"{int(progress * 100)}%")
                    
                    if video.get('video_id'):
                        try:
                            add_log(f"💬 댓글 수집 시도 - {video.get('title', 'Unknown')[:30]}... (ID: {video['video_id']})", "info")
                            comments = crawler.get_video_comments(video['video_id'], comments_per_video)
                            
                            if comments:
                                all_comments.extend(comments)
                                add_log(f"✅ 댓글 수집 성공 - {len(comments)}개 댓글 수집", "success")
                            else:
                                add_log(f"⚠️ 댓글 없음 - {video.get('title', 'Unknown')[:30]}...", "warning")
                                
                        except Exception as comment_error:
                            error_msg = str(comment_error)
                            add_log(f"❌ 댓글 수집 실패 - {video.get('title', 'Unknown')[:30]}... (오류: {error_msg[:100]}...)", "error")
                            
                            # 상세 오류 정보 표시
                            with st.expander(f"🔧 댓글 수집 오류 상세 정보 - {video.get('title', 'Unknown')[:30]}..."):
                                st.error(f"**오류 유형**: {type(comment_error).__name__}")
                                st.error(f"**오류 메시지**: {error_msg}")
                                st.error(f"**영상 ID**: {video.get('video_id', 'N/A')}")
                                st.error(f"**영상 제목**: {video.get('title', 'N/A')}")
                            
                            # ChromeDriver 재연결 시도
                            if "connection" in error_msg.lower() or "webdriver" in error_msg.lower():
                                try:
                                    add_log("🔄 ChromeDriver 재연결 시도 중...", "info")
                                    crawler.close()
                                    time.sleep(2)
                                    crawler = YouTubeCrawler()
                                    crawler.update_config(config)
                                    add_log("✅ ChromeDriver 재연결 성공", "success")
                                except Exception as reconnect_error:
                                    add_log(f"❌ ChromeDriver 재연결 실패: {str(reconnect_error)}", "error")
                                    break
            
            # 단계 4: 데이터 저장 (콤팩트)
            status_text.text("💾 데이터 저장 중...")
            progress_text.text("95%")
            add_log("💾 데이터 저장 시작", "info")
            progress_bar.progress(0.95)
            
            try:
                # Streamlit Cloud 환경에서 직접 엑셀 생성
                import io
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # 메모리에서 엑셀 파일 생성
                wb = Workbook()
                
                # 영상 정보 시트
                if videos:
                    ws_videos = wb.active
                    ws_videos.title = "Videos"
                    
                    # 헤더 추가
                    if videos:
                        headers = list(videos[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws_videos.cell(row=1, column=col, value=header)
                        
                        # 데이터 추가
                        for row, video in enumerate(videos, 2):
                            for col, header in enumerate(headers, 1):
                                ws_videos.cell(row=row, column=col, value=str(video.get(header, '')))
                
                # 댓글 정보 시트
                if all_comments:
                    ws_comments = wb.create_sheet("Comments")
                    
                    # 헤더 추가 (Comment 옆에 키워드 컬럼 추가)
                    if all_comments:
                        # 기존 헤더 가져오기
                        original_headers = list(all_comments[0].keys())
                        
                        # Comment 컬럼 위치 찾기
                        comment_col_index = None
                        for i, header in enumerate(original_headers):
                            if header == 'comment':
                                comment_col_index = i
                                break
                        
                        # 새로운 헤더 생성 (Comment 옆에 키워드 컬럼 추가)
                        new_headers = []
                        for i, header in enumerate(original_headers):
                            new_headers.append(header)
                            # Comment 컬럼 다음에 키워드 컬럼 추가
                            if i == comment_col_index:
                                new_headers.append('추출된_키워드(5개)')
                        
                        # 헤더 추가
                        for col, header in enumerate(new_headers, 1):
                            ws_comments.cell(row=1, column=col, value=header)
                        
                        # 데이터 추가
                        for row, comment in enumerate(all_comments, 2):
                            col_offset = 0
                            for i, header in enumerate(original_headers):
                                current_col = i + 1 + col_offset
                                ws_comments.cell(row=row, column=current_col, value=str(comment.get(header, '')))
                                
                                # Comment 컬럼 다음에 키워드 데이터 추가
                                if i == comment_col_index:
                                    col_offset += 1
                                    keyword_col = current_col + 1
                                    extracted_keywords = comment.get('extracted_keywords', '')
                                    # 키워드가 5개를 초과하면 5개만 표시
                                    if extracted_keywords:
                                        keywords_list = [kw.strip() for kw in extracted_keywords.split(',')]
                                        keywords_list = keywords_list[:5]  # 최대 5개만
                                        formatted_keywords = ', '.join(keywords_list)
                                        ws_comments.cell(row=row, column=keyword_col, value=formatted_keywords)
                                    else:
                                        ws_comments.cell(row=row, column=keyword_col, value='')
                
                # 키워드 분석 시트 (댓글이 있고 키워드 분석이 활성화된 경우)
                if all_comments and enable_keyword_analysis:
                    try:
                        from youtube_crawler import KeywordAnalyzer
                        
                        # 댓글 텍스트 추출 (comment 필드에서 추출)
                        comment_texts = [comment.get('comment', '') for comment in all_comments if comment.get('comment')]
                        
                        if comment_texts:
                            # 키워드 분석 수행 (Java 환경 문제 처리)
                            try:
                                analyzer = KeywordAnalyzer()
                                analysis_result = analyzer.analyze_keywords(comment_texts, top_n=20)
                            except Exception as java_error:
                                if "JVM" in str(java_error) or "Java" in str(java_error):
                                    st.warning("⚠️ Java 환경 문제로 고급 키워드 분석을 건너뜁니다. 기본 키워드 추출은 계속 작동합니다.")
                                    analysis_result = None
                                else:
                                    raise java_error
                            
                            if analysis_result:
                                # 키워드 분석 시트 생성
                                ws_keywords = wb.create_sheet("Keyword_Analysis")
                                
                                # 키워드 빈도 데이터
                                ws_keywords.cell(row=1, column=1, value="키워드")
                                ws_keywords.cell(row=1, column=2, value="빈도")
                                
                                row = 2
                                for keyword, freq in analysis_result.get('top_keywords', {}).items():
                                    ws_keywords.cell(row=row, column=1, value=keyword)
                                    ws_keywords.cell(row=row, column=2, value=freq)
                                    row += 1
                                
                                # 통계 정보
                                ws_keywords.cell(row=row+1, column=1, value="총 단어 수")
                                ws_keywords.cell(row=row+1, column=2, value=analysis_result.get('total_words', 0))
                                
                                ws_keywords.cell(row=row+2, column=1, value="고유 단어 수")
                                ws_keywords.cell(row=row+2, column=2, value=analysis_result.get('unique_words', 0))
                                
                                ws_keywords.cell(row=row+3, column=1, value="감정 점수")
                                ws_keywords.cell(row=row+3, column=2, value=analysis_result.get('sentiment_score', 0))
                                
                                ws_keywords.cell(row=row+4, column=1, value="감정 분류")
                                ws_keywords.cell(row=row+4, column=2, value=analysis_result.get('sentiment_label', '중립적'))
                                
                                # 성능 메트릭 시트
                                ws_metrics = wb.create_sheet("Performance_Metrics")
                                ws_metrics.cell(row=1, column=1, value="메트릭")
                                ws_metrics.cell(row=1, column=2, value="값")
                                
                                ws_metrics.cell(row=2, column=1, value="수집된 영상 수")
                                ws_metrics.cell(row=2, column=2, value=len(videos))
                                
                                ws_metrics.cell(row=3, column=1, value="수집된 댓글 수")
                                ws_metrics.cell(row=3, column=2, value=len(all_comments))
                                
                                ws_metrics.cell(row=4, column=1, value="분석된 댓글 수")
                                ws_metrics.cell(row=4, column=2, value=len(comment_texts))
                                
                                ws_metrics.cell(row=5, column=1, value="키워드 분석 완료")
                                ws_metrics.cell(row=5, column=2, value="성공")
                                
                    except Exception as analysis_error:
                        # 키워드 분석 실패 시 오류 정보 저장
                        ws_analysis_error = wb.create_sheet("Analysis_Error")
                        ws_analysis_error.cell(row=1, column=1, value="오류 유형")
                        ws_analysis_error.cell(row=1, column=2, value="오류 메시지")
                        ws_analysis_error.cell(row=2, column=1, value="키워드 분석 실패")
                        ws_analysis_error.cell(row=2, column=2, value=str(analysis_error))
                
                # 메모리에 엑셀 파일 저장
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                progress_bar.progress(1.0)
                status_text.text("✅ 크롤링 완료!")
                progress_text.text("100%")
                
                # 댓글 수집 결과 로그
                if collect_comments:
                    add_log(f"📊 댓글 수집 완료 - 총 {len(all_comments)}개 댓글 수집", "success")
                
                st.success(f"🎉 크롤링이 완료되었습니다!")
                st.info(f"📊 수집 결과: 영상 {len(videos)}개, 댓글 {len(all_comments)}개")
                
                # 세션 상태에 데이터 저장 (파일 다운로드용)
                st.session_state.videos = videos
                st.session_state.comments = all_comments
                st.session_state.excel_buffer = excel_buffer.getvalue()
                st.session_state.filename = filename
                st.session_state.crawling_completed = True
                
                # 디버깅 정보 표시
                add_log(f"💾 세션 상태 저장 완료 - 영상: {len(videos)}, 댓글: {len(all_comments)}", "info")
                
                # 성능 요약 표시
                performance_summary = performance_monitor.get_performance_summary()
                with st.expander("📊 성능 요약"):
                    st.write(f"**총 실행 시간**: {performance_summary.get('total_time', 0):.2f}초")
                    st.write(f"**평균 메모리 사용량**: {performance_summary.get('avg_memory_mb', 0):.1f} MB")
                    st.write(f"**평균 CPU 사용량**: {performance_summary.get('avg_cpu_percent', 0):.1f}%")
                    
                    if performance_summary.get('operations'):
                        st.write("**작업별 실행 시간**:")
                        for op, times in performance_summary['operations'].items():
                            avg_time = sum(times) / len(times)
                            st.write(f"- {op}: {avg_time:.2f}초 (평균)")
                
                # 결과를 캐시에 저장
                cache_result = {
                    'videos': videos,
                    'comments': all_comments,
                    'timestamp': datetime.now().isoformat()
                }
                cache_manager.save_to_cache(cache_key, cache_result)
                add_log("💾 결과를 캐시에 저장했습니다.", "info")
                
            except Exception as excel_error:
                st.error(f"❌ 엑셀 파일 생성 오류: {str(excel_error)}")
                # CSV로 대체
                if videos:
                    videos_df = pd.DataFrame(videos)
                    csv_videos = videos_df.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📥 영상 데이터 CSV 다운로드",
                        data=csv_videos,
                        file_name="videos.csv",
                        mime="text/csv"
                    )
                
                if all_comments:
                    comments_df = pd.DataFrame(all_comments)
                    csv_comments = comments_df.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📥 댓글 데이터 CSV 다운로드",
                        data=csv_comments,
                        file_name="comments.csv",
                        mime="text/csv"
                    )
                
                # 세션 상태에 데이터 저장
                st.session_state.videos = videos
                st.session_state.comments = all_comments
                st.session_state.filename = filename
                st.session_state.start_date = start_date
                st.session_state.end_date = end_date
                
                # 크롤링 완료 상태 저장
                st.session_state.crawling_completed = True
                
        except Exception as e:
            error_msg = str(e)
            add_log(f"❌ 크롤링 중 오류 발생: {error_msg[:100]}...", "error")
            
            # ChromeDriver 관련 오류인지 확인
            if "chromedriver" in error_msg.lower() or "webdriver" in error_msg.lower():
                st.error("""
                ❌ ChromeDriver 오류가 발생했습니다.
                
                **해결 방법:**
                1. Chrome 브라우저가 설치되어 있는지 확인
                2. ChromeDriver가 올바르게 설치되어 있는지 확인
                3. 브라우저를 다시 시작해보세요
                
                **상세 오류:** """ + error_msg)
            else:
                st.error(f"❌ 오류가 발생했습니다: {error_msg}")
            
            # 디버깅 정보 표시
            with st.expander("🔧 디버깅 정보"):
                st.code(f"""
                오류 유형: {type(e).__name__}
                오류 메시지: {error_msg}
                ChromeDriver 경로: /opt/homebrew/bin/chromedriver
                """)
        finally:
            if crawler:
                try:
                    crawler.close()
                except:
                    pass
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # 구분선
    st.markdown("---")
    
    # 결과 미리보기 및 파일 다운로드 영역
    if hasattr(st.session_state, 'crawling_completed') and st.session_state.crawling_completed:
        st.markdown('<h2 style="color: #1a202c; font-size: 1.5rem; font-weight: 600; margin-bottom: 1rem;">📊 크롤링 결과 미리보기</h2>', unsafe_allow_html=True)
        
        videos = st.session_state.get('videos', [])
        comments = st.session_state.get('comments', [])
        
        # 상단 메트릭 카드들
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div style="text-align: center;">
                    <h3 style="color: #FF6B6B; font-size: 2rem; margin: 0;">{len(videos)}</h3>
                    <p style="color: #666; margin: 0;">수집된 영상</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div style="text-align: center;">
                    <h3 style="color: #4ECDC4; font-size: 2rem; margin: 0;">{len(comments)}</h3>
                    <p style="color: #666; margin: 0;">수집된 댓글</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            unique_channels = len(set([video.get('channel_name', 'Unknown') for video in videos]))
            st.markdown(f"""
            <div class="metric-card">
                <div style="text-align: center;">
                    <h3 style="color: #FFD93D; font-size: 2rem; margin: 0;">{unique_channels}</h3>
                    <p style="color: #666; margin: 0;">채널 수</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div class="metric-card">
                <div style="text-align: center;">
                    <h3 style="color: #A8E6CF; font-size: 2rem; margin: 0;">📅</h3>
                    <p style="color: #666; margin: 0;">수집 완료</p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # 데이터 미리보기와 파일 다운로드를 나란히 배치
        col_preview, col_download = st.columns([2, 1])
        
        with col_preview:
            st.markdown('<h3 style="color: #1a202c; font-size: 1.3rem; font-weight: 600; margin-bottom: 1rem;">📋 데이터 미리보기</h3>', unsafe_allow_html=True)
            
            # 탭 생성
            tab1, tab2 = st.tabs(["🎥 영상 목록", "💬 댓글 목록"])
            
            with tab1:
                if videos:
                    df_videos = pd.DataFrame(videos)
                    st.dataframe(df_videos.head(10), use_container_width=True)  # 상위 10개만 표시
                    if len(videos) > 10:
                        st.info(f"📊 총 {len(videos)}개 영상 중 상위 10개를 표시합니다.")
                else:
                    st.info("🎥 수집된 영상이 없습니다.")
            
            with tab2:
                if comments and len(comments) > 0:
                    df_comments = pd.DataFrame(comments)
                    st.dataframe(df_comments.head(10), use_container_width=True)  # 상위 10개만 표시
                    if len(comments) > 10:
                        st.info(f"📊 총 {len(comments)}개 댓글 중 상위 10개를 표시합니다.")
                    
                    # 댓글 데이터 디버깅 정보
                    with st.expander("🔧 댓글 데이터 디버깅 정보"):
                        st.write(f"**댓글 개수**: {len(comments)}개")
                        st.write(f"**댓글 컬럼**: {list(df_comments.columns)}")
                        if len(comments) > 0:
                            st.write(f"**첫 번째 댓글 샘플**:")
                            st.json(comments[0])
                else:
                    st.warning("💬 수집된 댓글이 없습니다.")
                    st.info("💡 댓글 수집이 비활성화되었거나 댓글 수집에 실패했을 수 있습니다.")
        
        with col_download:
            st.markdown('<h3 style="color: #1a202c; font-size: 1.3rem; font-weight: 600; margin-bottom: 1rem;">📥 파일 다운로드</h3>', unsafe_allow_html=True)
            
            # 파일 형식 선택
            file_format = st.selectbox(
                "파일 형식 선택",
                options=["XLSX (Excel)", "CSV"],
                help="다운로드할 파일 형식을 선택하세요"
            )
            
            # 새로고침 버튼
            if st.button("🔄 새로고침", help="페이지를 새로고침합니다"):
                st.rerun()
            
            # 파일 다운로드 버튼들
            if file_format == "XLSX (Excel)" and hasattr(st.session_state, 'excel_buffer'):
                excel_data = st.session_state.excel_buffer
                filename = st.session_state.get('filename', 'youtube_data.xlsx')
                
                # 히스토리에 다운로드 기록 추가
                download_record = history_manager.add_download_record(
                    filename=filename,
                    data_type="Excel",
                    record_count=len(videos) + len(comments),
                    file_size=len(excel_data),
                    download_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
                
                st.download_button(
                    label="📥 엑셀 파일 다운로드",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="수집된 모든 데이터가 포함된 엑셀 파일을 다운로드합니다"
                )
            
            elif file_format == "CSV":
                if videos:
                    videos_df = pd.DataFrame(videos)
                    csv_videos = videos_df.to_csv(index=False, encoding='utf-8-sig')
                    
                    # 히스토리에 다운로드 기록 추가
                    history_manager.add_download_record(
                        filename="videos.csv",
                        data_type="CSV (Videos)",
                        record_count=len(videos),
                        file_size=len(csv_videos.encode('utf-8-sig')),
                        download_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    )
                    
                    st.download_button(
                        label="📥 영상 데이터 CSV",
                        data=csv_videos,
                        file_name="videos.csv",
                        mime="text/csv",
                        help="영상 데이터만 포함된 CSV 파일을 다운로드합니다"
                    )
                
                if comments:
                    comments_df = pd.DataFrame(comments)
                    csv_comments = comments_df.to_csv(index=False, encoding='utf-8-sig')
                    
                    # 히스토리에 다운로드 기록 추가
                    history_manager.add_download_record(
                        filename="comments.csv",
                        data_type="CSV (Comments)",
                        record_count=len(comments),
                        file_size=len(csv_comments.encode('utf-8-sig')),
                        download_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    )
                    
                    st.download_button(
                        label="📥 댓글 데이터 CSV",
                        data=csv_comments,
                        file_name="comments.csv",
                        mime="text/csv",
                        help="댓글 데이터만 포함된 CSV 파일을 다운로드합니다"
                    )
            
            # 히스토리 관리 섹션
            st.markdown("---")
            st.markdown("### 📋 다운로드 히스토리")
            
            recent_history = history_manager.get_recent_history(5)
            if recent_history:
                for record in recent_history:
                    with st.expander(f"📄 {record['filename']} ({record['download_time']})"):
                        st.write(f"**파일 유형**: {record['data_type']}")
                        st.write(f"**레코드 수**: {record['record_count']:,}개")
                        st.write(f"**파일 크기**: {record['file_size_mb']} MB")
                        st.write(f"**다운로드 시간**: {record['download_time']}")
            else:
                st.info("📋 다운로드 히스토리가 없습니다.")
            
            # 히스토리 관리 버튼들
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🗑️ 데이터 초기화", help="수집된 데이터를 모두 삭제합니다"):
                    for key in ['videos', 'comments', 'excel_buffer', 'filename', 'crawling_completed']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()
            
            with col2:
                if st.button("🗑️ 히스토리 삭제", help="다운로드 히스토리를 모두 삭제합니다"):
                    history_manager.clear_history()
                    st.success("히스토리가 삭제되었습니다!")
                    st.rerun()
    
    # 데이터가 없을 때 안내 메시지
    else:
        st.markdown("""
        <div style="text-align: center; padding: 3rem; background: #f8fafc; border-radius: 10px; margin: 2rem 0;">
            <h3 style="color: #4a5568; margin-bottom: 1rem;">🚀 크롤링을 시작해보세요!</h3>
            <p style="color: #666; margin-bottom: 2rem;">위의 설정을 완료하고 크롤링 시작 버튼을 눌러 데이터를 수집하세요.</p>
            <div style="display: flex; justify-content: center; gap: 1rem;">
                <div style="background: #4ECDC4; color: white; padding: 0.5rem 1rem; border-radius: 5px; font-size: 0.9rem;">🔍 영상 검색</div>
                <div style="background: #FF6B6B; color: white; padding: 0.5rem 1rem; border-radius: 5px; font-size: 0.9rem;">💬 댓글 수집</div>
                <div style="background: #FFD93D; color: white; padding: 0.5rem 1rem; border-radius: 5px; font-size: 0.9rem;">📊 데이터 분석</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # 두 번째 탭: 영상 ID 댓글 추출
    with tab2:
        st.markdown('<h2 style="color: #1a202c; font-size: 1.5rem; font-weight: 600; margin-bottom: 1rem;">💬 영상 ID 댓글 추출</h2>', unsafe_allow_html=True)
        
        # 영상 ID 입력 섹션
        st.markdown("### 📝 영상 ID 입력")
        st.info("💡 **영상 ID 찾는 방법**:\n"
               "1. 유튜브 영상 URL에서 `v=` 뒤의 11자리 코드\n"
               "2. 예시: `https://www.youtube.com/watch?v=dQw4w9WgXcQ` → `dQw4w9WgXcQ`\n"
               "3. 여러 영상의 댓글을 한 번에 수집할 수 있습니다")
        
        # 영상 ID 입력 방식 선택
        input_method = st.radio(
            "입력 방식 선택",
            ["단일 영상 ID", "여러 영상 ID (한 줄에 하나씩)"],
            help="하나의 영상 ID를 입력하거나 여러 영상 ID를 한 번에 입력할 수 있습니다"
        )
        
        if input_method == "단일 영상 ID":
            video_id = st.text_input(
                "영상 ID",
                placeholder="예: dQw4w9WgXcQ",
                help="추출할 댓글의 영상 ID를 입력하세요"
            )
            video_ids = [video_id.strip()] if video_id.strip() else []
        else:
            video_ids_text = st.text_area(
                "영상 ID 목록",
                placeholder="dQw4w9WgXcQ\njNQXAC9IVRw\n...",
                height=150,
                help="한 줄에 하나씩 영상 ID를 입력하세요"
            )
            video_ids = [vid.strip() for vid in video_ids_text.split('\n') if vid.strip()]
        
        # 댓글 수집 설정
        st.markdown("### ⚙️ 댓글 수집 설정")
        col1, col2 = st.columns(2)
        
        with col1:
            comments_per_video = st.number_input(
                "영상당 댓글 수",
                min_value=1, max_value=100, value=20,
                step=1,
                help="각 영상에서 수집할 댓글의 수"
            )
        
        with col2:
            enable_keyword_analysis = st.checkbox(
                "댓글 키워드 분석",
                value=True,
                help="댓글에서 키워드 및 감정 분석 수행"
            )
        
        # 크롤링 시작 버튼
        if st.button("🚀 댓글 추출 시작", type="primary", use_container_width=True):
            if not video_ids:
                st.error("❌ 영상 ID를 입력해주세요.")
            else:
                # 댓글 추출 진행
                with st.spinner("🔄 댓글 추출 중..."):
                    try:
                        # 크롤러 초기화
                        crawler = YouTubeCrawler()
                        
                        # 설정 적용
                        config = {
                            'max_workers': 2,  # 댓글 수집은 적은 워커 사용
                            'enable_keyword_analysis': enable_keyword_analysis,
                            'excel_encoding': 'utf-8-sig',
                            'max_comments_per_video': comments_per_video,
                            'comment_batch_size': 10
                        }
                        crawler.update_config(config)
                        
                        # 댓글 수집
                        all_comments = []
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        for i, video_id in enumerate(video_ids):
                            status_text.text(f"📹 영상 {i+1}/{len(video_ids)} 처리 중: {video_id}")
                            
                            try:
                                comments = crawler.get_video_comments(video_id, comments_per_video)
                                if comments:
                                    # 영상 ID 정보 추가
                                    for comment in comments:
                                        comment['video_id'] = video_id
                                    all_comments.extend(comments)
                                    st.success(f"✅ {video_id}: {len(comments)}개 댓글 수집 완료")
                                else:
                                    st.warning(f"⚠️ {video_id}: 댓글을 찾을 수 없습니다")
                            except Exception as e:
                                st.error(f"❌ {video_id}: 오류 발생 - {str(e)}")
                            
                            progress_bar.progress((i + 1) / len(video_ids))
                        
                        crawler.close()
                        
                        # 결과 저장
                        if all_comments:
                            st.session_state.comments_only = all_comments
                            st.session_state.comments_extraction_completed = True
                            st.session_state.video_ids_processed = video_ids
                            st.session_state.keyword_analysis_enabled = enable_keyword_analysis
                            
                            st.success(f"🎉 댓글 추출 완료! 총 {len(all_comments)}개 댓글을 수집했습니다.")
                            st.rerun()
                        else:
                            st.error("❌ 수집된 댓글이 없습니다.")
                    
                    except Exception as e:
                        st.error(f"❌ 댓글 추출 중 오류 발생: {str(e)}")
        
        # 댓글 추출 결과 표시
        if hasattr(st.session_state, 'comments_extraction_completed') and st.session_state.comments_extraction_completed:
            st.markdown("---")
            st.markdown('<h3 style="color: #1a202c; font-size: 1.3rem; font-weight: 600; margin-bottom: 1rem;">📊 댓글 추출 결과</h3>', unsafe_allow_html=True)
            
            comments = st.session_state.get('comments_only', [])
            video_ids_processed = st.session_state.get('video_ids_processed', [])
            
            # 결과 통계
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("처리된 영상", len(video_ids_processed))
            with col2:
                st.metric("수집된 댓글", len(comments))
            with col3:
                avg_comments = len(comments) / len(video_ids_processed) if video_ids_processed else 0
                st.metric("평균 댓글/영상", f"{avg_comments:.1f}")
            
            # 댓글 데이터 표시
            if comments:
                df_comments = pd.DataFrame(comments)
                
                # 탭으로 데이터와 분석 분리
                tab_data, tab_analysis = st.tabs(["📋 댓글 데이터", "🔍 키워드 분석"])
                
                with tab_data:
                    st.dataframe(df_comments, use_container_width=True)
                
                with tab_analysis:
                    # 키워드 분석 수행
                    # 세션에서 키워드 분석 설정 가져오기
                    keyword_analysis_enabled = st.session_state.get('keyword_analysis_enabled', True)
                    if keyword_analysis_enabled:
                        st.markdown("### 🔍 키워드 분석 결과")
                        
                        # 분석 진행 표시
                        with st.spinner("🔍 키워드 분석 중..."):
                            try:
                                # 모든 댓글 텍스트 수집
                                all_texts = []
                                for comment in comments:
                                    text = comment.get('text', '')
                                    if text and isinstance(text, str):
                                        all_texts.append(text)
                                
                                if all_texts:
                                    # 키워드 분석 함수 호출
                                    keyword_results = perform_keyword_analysis(all_texts)
                                    
                                    # 분석 결과 표시
                                    col1, col2 = st.columns(2)
                                    
                                    with col1:
                                        st.markdown("#### 📊 상위 키워드")
                                        if keyword_results['top_keywords']:
                                            for i, (keyword, count) in enumerate(keyword_results['top_keywords'][:10], 1):
                                                st.markdown(f"**{i}.** {keyword} ({count}회)")
                                        else:
                                            st.info("분석된 키워드가 없습니다.")
                                    
                                    with col2:
                                        st.markdown("#### 📈 키워드 분포")
                                        if keyword_results['keyword_stats']:
                                            st.write(f"**총 키워드 수**: {keyword_results['keyword_stats']['total_keywords']}")
                                            st.write(f"**고유 키워드 수**: {keyword_results['keyword_stats']['unique_keywords']}")
                                            st.write(f"**평균 키워드 길이**: {keyword_results['keyword_stats']['avg_length']:.1f}")
                                        else:
                                            st.info("키워드 통계를 계산할 수 없습니다.")
                                    
                                    # 감정 분석 결과
                                    if keyword_results['sentiment_analysis']:
                                        st.markdown("#### 😊 감정 분석")
                                        sentiment = keyword_results['sentiment_analysis']
                                        
                                        col1, col2, col3 = st.columns(3)
                                        with col1:
                                            st.metric("긍정적", f"{sentiment['positive']:.1f}%")
                                        with col2:
                                            st.metric("중립적", f"{sentiment['neutral']:.1f}%")
                                        with col3:
                                            st.metric("부정적", f"{sentiment['negative']:.1f}%")
                                    
                                    # 영상별 키워드 분석
                                    st.markdown("#### 🎥 영상별 키워드 분석")
                                    video_keywords = {}
                                    
                                    for video_id in video_ids_processed:
                                        video_comments = [c for c in comments if c.get('video_id') == video_id]
                                        if video_comments:
                                            video_texts = [c.get('text', '') for c in video_comments if c.get('text')]
                                            if video_texts:
                                                video_analysis = perform_keyword_analysis(video_texts)
                                                video_keywords[video_id] = video_analysis
                                    
                                    # 영상별 키워드 표시
                                    for video_id, analysis in video_keywords.items():
                                        with st.expander(f"📹 영상 {video_id}"):
                                            if analysis['top_keywords']:
                                                st.write("**상위 키워드**:")
                                                for keyword, count in analysis['top_keywords'][:5]:
                                                    st.write(f"- {keyword} ({count}회)")
                                            else:
                                                st.info("이 영상에서 분석된 키워드가 없습니다.")
                                    
                                    # 분석 결과를 세션에 저장
                                    st.session_state.keyword_analysis_results = keyword_results
                                    st.session_state.video_keywords_analysis = video_keywords
                                    
                                else:
                                    st.warning("⚠️ 분석할 댓글 텍스트가 없습니다.")
                                    
                            except Exception as e:
                                st.error(f"❌ 키워드 분석 중 오류 발생: {str(e)}")
                    else:
                        st.info("💡 키워드 분석이 비활성화되어 있습니다. 설정에서 활성화하세요.")
                
                # 파일 다운로드
                st.markdown("### 📥 댓글 데이터 다운로드")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    # CSV 다운로드
                    csv_data = df_comments.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📥 CSV 다운로드",
                        data=csv_data,
                        file_name=f"comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                
                with col2:
                    # 엑셀 다운로드 (댓글만)
                    from io import BytesIO
                    with BytesIO() as buffer:
                        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                            df_comments.to_excel(writer, sheet_name='Comments', index=False)
                        excel_data = buffer.getvalue()
                    
                    st.download_button(
                        label="📥 Excel 다운로드",
                        data=excel_data,
                        file_name=f"comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col3:
                    # 키워드 분석 결과가 있는 경우 분석 결과 포함 엑셀 다운로드
                    if hasattr(st.session_state, 'keyword_analysis_results') and st.session_state.keyword_analysis_results:
                        keyword_results = st.session_state.keyword_analysis_results
                        
                        # 키워드 분석 결과를 DataFrame으로 변환
                        keyword_df = pd.DataFrame(keyword_results['top_keywords'], columns=['키워드', '빈도'])
                        
                        # 감정 분석 결과를 DataFrame으로 변환
                        sentiment_df = pd.DataFrame([keyword_results['sentiment_analysis']])
                        
                        # 통계 정보를 DataFrame으로 변환
                        stats_df = pd.DataFrame([keyword_results['keyword_stats']])
                        
                        with BytesIO() as buffer:
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                df_comments.to_excel(writer, sheet_name='Comments', index=False)
                                keyword_df.to_excel(writer, sheet_name='Keywords', index=False)
                                sentiment_df.to_excel(writer, sheet_name='Sentiment', index=False)
                                stats_df.to_excel(writer, sheet_name='Statistics', index=False)
                            analysis_excel_data = buffer.getvalue()
                        
                        st.download_button(
                            label="📊 분석 포함 Excel",
                            data=analysis_excel_data,
                            file_name=f"comments_with_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.info("💡 키워드 분석을 활성화하면 분석 결과가 포함된 파일을 다운로드할 수 있습니다.")
                
                # 데이터 초기화
                if st.button("🗑️ 댓글 데이터 초기화"):
                    for key in ['comments_only', 'comments_extraction_completed', 'video_ids_processed', 'keyword_analysis_enabled', 'keyword_analysis_results', 'video_keywords_analysis']:
                        if key in st.session_state:
                            del st.session_state[key]
                    st.rerun()

if __name__ == "__main__":
    main() 